import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

# Dados da tabela principal
data = {
    "Categoria": ["TV", "Armário", "Liquidificador", "Celular", "Fone", "Microondas", "Cama", "Edredon"],
    "Preço Unitário": [1200, 450, 80, 1500, 20, 375, 600, 148],
    "Quantidade": [1, 2, 2, 4, 3, 5, 2, 6],
    "Total": ["=B2*C2", "=B3*C3", "=B4*C4", "=B5*C5", "=B6*C6", "=B7*C7", "=B8*C8", "=B9*C9"],
    "Vendedor": ["Junior", "Cesar", "Junior", "Patricia", "Magali", "Junior", "Bruna", "Magali"]
}

df = pd.DataFrame(data)

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Controle de Vendas"

# Estilos
rosa = PatternFill(start_color="F4B6C2", end_color="F4B6C2", fill_type="solid")
negrito = Font(bold=True, color="FFFFFF")
centro = Alignment(horizontal="center", vertical="center")
borda = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))

# Cabeçalho principal
ws.merge_cells("A1:E1")
titulo = ws["A1"]
titulo.value = "Controle de Vendas"
titulo.fill = rosa
titulo.font = Font(bold=True, color="FFFFFF", size=14)
titulo.alignment = centro

# Cabeçalhos
headers = list(df.columns)
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.fill = rosa
    cell.font = negrito
    cell.alignment = centro
    cell.border = borda

# Inserir dados
for i, row in df.iterrows():
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i+3, column=j, value=val)
        cell.border = borda
        if j in [2, 4]:  # Colunas de preço e total
            cell.number_format = 'R$ #,##0.00'
        cell.alignment = Alignment(horizontal="center")

# Ajustar larguras
larguras = [18, 15, 12, 15, 15]
for i, largura in enumerate(larguras, start=1):
    ws.column_dimensions[chr(64+i)].width = largura

# Salvar arquivo Excel
wb.save("Controle_de_Vendas.xlsx")
print("✅ Arquivo 'Controle_de_Vendas.xlsx' criado com sucesso!")

